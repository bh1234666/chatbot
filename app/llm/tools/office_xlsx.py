"""xlsx 电子表格操作:读/写/追加/单元格更新/多表应用/完整性校验/列字母换算 + 返回上限常量。

2026-05-20 重构: 从 llm/tools/office.py 原样抽出。closure 自包含(10 符号, 0 unsafe);
_err/_jsonable 复用 office_common。openpyxl 在函数内惰性 import。office.py re-export 兼容。
"""
from __future__ import annotations

import asyncio
import json
import os
from typing import Any

from app.llm.tools.office_common import _err, _jsonable


_MAX_SHEETS_RETURNED = 50


_MAX_XLSX_ROWS_RETURNED = 500


_MAX_XLSX_COLS_RETURNED = 50


async def _xlsx_read(workspace_dir: str, target: str, rel_path: str, args: dict) -> str:
    if not os.path.isfile(target):
        return _err(f"file not found: {rel_path}")

    import openpyxl

    def _do_read():
        try:
            wb = openpyxl.load_workbook(target, data_only=True, read_only=True)
        except Exception as e:
            return None, f"cannot open as xlsx ({type(e).__name__}: {e})"

        sheets_out: list[dict] = []
        truncated_sheets = False
        sheetnames = list(wb.sheetnames)
        for s_idx, sheet_name in enumerate(sheetnames):
            if s_idx >= _MAX_SHEETS_RETURNED:
                truncated_sheets = True
                break
            ws = wb[sheet_name]
            rows: list[list] = []
            truncated_rows = False
            truncated_cols = False
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx >= _MAX_XLSX_ROWS_RETURNED:
                    truncated_rows = True
                    break
                cells = []
                for c_idx, val in enumerate(row):
                    if c_idx >= _MAX_XLSX_COLS_RETURNED:
                        truncated_cols = True
                        break
                    cells.append(_jsonable(val))
                rows.append(cells)
            sheets_out.append({
                "name": sheet_name,
                "row_count": ws.max_row or 0,
                "col_count": ws.max_column or 0,
                "rows": rows,
                **({"truncated_rows": True} if truncated_rows else {}),
                **({"truncated_cols": True} if truncated_cols else {}),
            })
        wb.close()
        return {
            "sheetnames": sheetnames,
            "sheets_out": sheets_out,
            "truncated_sheets": truncated_sheets,
        }, None

    result, err = await asyncio.to_thread(_do_read)
    if err:
        return _err(err)

    return json.dumps({
        "ok": True, "action": "read", "format": "xlsx", "path": rel_path,
        "sheet_count": len(result["sheetnames"]),
        "sheets": result["sheets_out"],
        **({"truncated_sheets": True} if result["truncated_sheets"] else {}),
    }, ensure_ascii=False)


# 2026-05-17 P162.7: xlsx 完整性检查
async def _xlsx_verify_integrity(
    workspace_dir: str, target: str, rel_path: str, args: dict,
) -> str:
    """对 .xlsx 做跨 sheet / 跨 column 的一致性 + 公式 vs 计算值校验。

    主要场景: 财务模型/实验数据表里同一指标在多 sheet/多 column 出现, 必须一致。

    args:
      tolerance: float = 0.0001     数值匹配相对容差
      key_columns: list[str] | None 用作行键的列名 (header 里查找)
      value_columns: list[str] | None 应该一致的指标列名
      check_formulas: bool = True   公式 cached_value 健全性
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    if not os.path.isfile(target):
        return _err(f"file not found: {rel_path}", action="verify_integrity")
    try:
        tolerance = float(args.get("tolerance", 0.0001))
    except (TypeError, ValueError):
        tolerance = 0.0001
    key_columns = args.get("key_columns") or []
    value_columns = args.get("value_columns") or []
    check_formulas = bool(args.get("check_formulas", True))

    def _do():
        try:
            wb_formula = openpyxl.load_workbook(target, data_only=False, read_only=False)
        except Exception as e:
            return None, f"open(formula) failed: {type(e).__name__}: {e}"
        try:
            wb_values = openpyxl.load_workbook(target, data_only=True, read_only=False)
        except Exception as e:
            wb_formula.close()
            return None, f"open(values) failed: {type(e).__name__}: {e}"

        cross_sheet_findings: list[dict] = []
        formula_anomalies: list[dict] = []

        # ── 1. 跨 sheet 一致性 ──
        if key_columns and value_columns:
            unified: dict[str, dict[str, dict[str, float]]] = {}
            for sheet_name in wb_values.sheetnames:
                ws = wb_values[sheet_name]
                # 找 header 行 (前 3 行)
                header_row = None
                header_row_idx = None
                for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
                    if any(str(v).strip() in key_columns for v in row if v is not None):
                        header_row = row
                        header_row_idx = r_idx
                        break
                if not header_row:
                    continue
                headers = [str(v).strip() if v is not None else "" for v in header_row]
                key_col_indices = {kc: headers.index(kc) for kc in key_columns if kc in headers}
                val_col_indices = {vc: headers.index(vc) for vc in value_columns if vc in headers}
                if not key_col_indices or not val_col_indices:
                    continue
                for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if r_idx <= header_row_idx:
                        continue
                    key_parts = []
                    for kc, ki in key_col_indices.items():
                        if ki < len(row) and row[ki] is not None:
                            key_parts.append(f"{kc}={row[ki]}")
                    if not key_parts:
                        continue
                    key_value = "|".join(key_parts)
                    for vc, vi in val_col_indices.items():
                        if vi < len(row):
                            v = row[vi]
                            try:
                                fv = float(v)
                                unified.setdefault(key_value, {}).setdefault(sheet_name, {})[vc] = fv
                            except (TypeError, ValueError):
                                pass
            for key_value, sheet_dict in unified.items():
                if len(sheet_dict) < 2:
                    continue
                for vc in value_columns:
                    sheet_to_val = {s: d[vc] for s, d in sheet_dict.items() if vc in d}
                    if len(sheet_to_val) < 2:
                        continue
                    vals = list(sheet_to_val.values())
                    vmax, vmin = max(vals), min(vals)
                    if vmax == 0:
                        continue
                    spread = (vmax - vmin) / abs(vmax)
                    if spread > tolerance:
                        cross_sheet_findings.append({
                            "key_value": key_value,
                            "value_col": vc,
                            "values_by_sheet": sheet_to_val,
                            "max_spread_pct": round(spread * 100, 3),
                        })

        # ── 2. 公式 vs cached value 健全性 ──
        if check_formulas:
            for sheet_name in wb_formula.sheetnames:
                ws_f = wb_formula[sheet_name]
                ws_v = wb_values[sheet_name]
                for row_idx in range(1, min(ws_f.max_row + 1, 501)):
                    for col_idx in range(1, min(ws_f.max_column + 1, 51)):
                        cell_f = ws_f.cell(row=row_idx, column=col_idx)
                        if not isinstance(cell_f.value, str):
                            continue
                        if not cell_f.value.startswith("="):
                            continue
                        cell_v = ws_v.cell(row=row_idx, column=col_idx)
                        cached = cell_v.value
                        coord = f"{get_column_letter(col_idx)}{row_idx}"
                        if cached is None:
                            formula_anomalies.append({
                                "sheet": sheet_name,
                                "cell": coord,
                                "formula": cell_f.value[:80],
                                "issue": "formula_no_cached_value",
                            })
                        elif isinstance(cached, str) and cached.startswith("="):
                            formula_anomalies.append({
                                "sheet": sheet_name,
                                "cell": coord,
                                "formula": cell_f.value[:80],
                                "cached_value": cached[:80],
                                "issue": "cached_value_still_formula_string",
                            })

        sheetnames = list(wb_values.sheetnames)
        wb_formula.close()
        wb_values.close()

        return {
            "cross_sheet_findings": cross_sheet_findings[:30],
            "cross_sheet_findings_count": len(cross_sheet_findings),
            "formula_anomalies": formula_anomalies[:30],
            "formula_anomalies_count": len(formula_anomalies),
            "sheets_checked": sheetnames,
        }, None

    result, err = await asyncio.to_thread(_do)
    if err:
        return _err(err, action="verify_integrity")

    n_cross = result["cross_sheet_findings_count"]
    n_form = result["formula_anomalies_count"]
    if n_cross == 0 and n_form == 0:
        hint = "XLSX integrity check passed: cross-sheet values are consistent and formulas have cached values.\nxlsx 完整性检查通过。"
    else:
        parts = []
        if n_cross > 0:
            parts.append(f"{n_cross} 处跨 sheet 数值不一致")
        if n_form > 0:
            parts.append(f"{n_form} 处公式单元格缺 cached value (Excel 没重算就保存)")
        hint = "⚠️ xlsx 完整性问题: " + "; ".join(parts)

    return json.dumps({
        "ok": True,
        "action": "verify_integrity",
        "format": "xlsx",
        "path": rel_path,
        "tolerance": tolerance,
        "hint": hint,
        **result,
    }, ensure_ascii=False)


async def _xlsx_write(workspace_dir: str, target: str, rel_path: str, args: dict) -> str:
    import openpyxl

    sheets = args.get("sheets") or []
    if not isinstance(sheets, list) or not sheets:
        return _err("sheets must be a non-empty array", action="write", hint="Pass sheets as [{name:'Sheet1', rows:[['A','B'],[1,2]], header:true}, ...]")

    def _do_write():
        wb = openpyxl.Workbook()
        # 删除默认 Sheet,后面按用户指定创建
        if wb.sheetnames:
            del wb[wb.sheetnames[0]]
        err = _apply_sheets_xlsx(wb, sheets)
        if err:
            return None, err
        if not wb.sheetnames:
            return None, "no sheets created (sheets array yielded nothing)"
        os.makedirs(os.path.dirname(target) or ".", exist_ok=True)
        try:
            wb.save(target)
        except Exception as e:
            return None, f"save failed: {type(e).__name__}: {e}"
        size = os.path.getsize(target) if os.path.exists(target) else 0
        return size, None

    size, err = await asyncio.to_thread(_do_write)
    if err:
        return _err(err)
    return json.dumps({
        "ok": True, "action": "write", "format": "xlsx", "path": rel_path,
        "sheets_written": len(sheets), "size_bytes": size,
    }, ensure_ascii=False)


async def _xlsx_append(workspace_dir: str, target: str, rel_path: str, args: dict) -> str:
    """xlsx append: 增加新 sheet 到已有工作簿。同名 sheet 会被替换。"""
    import openpyxl

    if not os.path.isfile(target):
        return _err(f"file not found: {rel_path}")

    sheets = args.get("sheets") or []
    if not isinstance(sheets, list) or not sheets:
        return _err("sheets must be a non-empty array")

    def _do_append():
        try:
            wb = openpyxl.load_workbook(target)
        except Exception as e:
            return f"open failed: {type(e).__name__}: {e}"
        err = _apply_sheets_xlsx(wb, sheets, replace_if_exists=True)
        if err:
            return err
        try:
            wb.save(target)
        except Exception as e:
            return f"save failed: {type(e).__name__}: {e}"
        return None

    err = await asyncio.to_thread(_do_append)
    if err:
        return _err(err)
    return json.dumps({
        "ok": True, "action": "append", "format": "xlsx", "path": rel_path,
        "sheets_added_or_replaced": len(sheets),
    }, ensure_ascii=False)


def _apply_sheets_xlsx(wb, sheets: list, *, replace_if_exists: bool = False) -> str | None:
    from openpyxl.styles import Font

    bold_font = Font(bold=True)

    for s_idx, s in enumerate(sheets):
        if not isinstance(s, dict):
            return f"sheets[{s_idx}] must be an object"
        name = str(s.get("name", f"Sheet{s_idx + 1}")).strip()[:31] or f"Sheet{s_idx + 1}"
        if name in wb.sheetnames:
            if replace_if_exists:
                del wb[name]
            else:
                base = name[:28]
                k = 2
                while f"{base}_{k}" in wb.sheetnames:
                    k += 1
                name = f"{base}_{k}"
        ws = wb.create_sheet(title=name)

        rows = s.get("rows") or []
        if not isinstance(rows, list):
            return f"sheets[{s_idx}].rows must be array"
        for r_idx, r in enumerate(rows):
            if not isinstance(r, list):
                continue
            for c_idx, val in enumerate(r):
                cell = ws.cell(row=r_idx + 1, column=c_idx + 1)
                # 字符串以 = 开头视为公式,openpyxl 自动识别
                cell.value = val if val is not None else ""

        if s.get("header") and rows:
            n_cols_hdr = len(rows[0]) if isinstance(rows[0], list) else 0
            for c_idx in range(n_cols_hdr):
                ws.cell(row=1, column=c_idx + 1).font = bold_font

        col_widths = s.get("column_widths")
        if isinstance(col_widths, list):
            for c_idx, w in enumerate(col_widths):
                if isinstance(w, (int, float)) and w > 0:
                    col_letter = _openpyxl_col_letter(c_idx + 1)
                    ws.column_dimensions[col_letter].width = float(w)

    return None


def _openpyxl_col_letter(n: int) -> str:
    """1 → A, 26 → Z, 27 → AA"""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


async def _xlsx_update_cells(
    workspace_dir: str, target: str, rel_path: str, args: dict,
) -> str:
    """更新一组单元格。

    args:
      updates: [{"sheet": "Sheet1", "ref": "B3", "value": 42}, ...]
              ref 用 Excel 风格的 A1 表示;value 可以是 string/number/bool/None
              sheet 可省略,默认第一张
      create_sheet_if_missing: bool,默认 False(sheet 不存在则报错)
    """
    import openpyxl

    if not os.path.isfile(target):
        return _err(f"file not found: {rel_path}", action="update_cells")
    updates = args.get("updates") or []
    if not isinstance(updates, list) or not updates:
        return _err("updates must be a non-empty array", action="update_cells", hint="Pass updates as [{sheet:'Sheet1', ref:'B3', value:42}, {sheet:'Sheet1', ref:'C5', value:'=SUM(C1:C4)'}, ...]")
    create_missing = bool(args.get("create_sheet_if_missing", False))

    def _do():
        try:
            wb = openpyxl.load_workbook(target)
        except Exception as e:
            return None, f"open failed: {type(e).__name__}: {e}"

        applied = 0
        skipped: list[str] = []
        for i, u in enumerate(updates):
            if not isinstance(u, dict):
                skipped.append(f"updates[{i}]: not a dict")
                continue
            sheet_name = u.get("sheet")
            ref = u.get("ref")
            if not ref or not isinstance(ref, str):
                skipped.append(f"updates[{i}]: missing/invalid ref")
                continue
            if sheet_name is None:
                ws = wb.active
            elif sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif create_missing:
                ws = wb.create_sheet(sheet_name)
            else:
                skipped.append(
                    f"updates[{i}]: sheet {sheet_name!r} not found "
                    f"(set create_sheet_if_missing=true to create)"
                )
                continue
            try:
                ws[ref] = u.get("value")
                applied += 1
            except Exception as e:
                skipped.append(f"updates[{i}] ref={ref}: {type(e).__name__}: {e}")

        try:
            wb.save(target)
        except Exception as e:
            return None, f"save failed: {type(e).__name__}: {e}"
        return {
            "cells_updated": applied,
            "skipped": skipped[:20],
            "total_requested": len(updates),
        }, None

    info, err = await asyncio.to_thread(_do)
    if err:
        return _err(err)
    return json.dumps({
        "ok": True, "action": "update_cells", "format": "xlsx",
        "path": rel_path,
        **info,
    }, ensure_ascii=False)
